from django.shortcuts import render
from django.http import HttpResponse
from django.forms import inlineformset_factory
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from openpyxl import Workbook
from django.http import Http404
from django.contrib.auth.models import Group
from django.views.defaults import page_not_found

from .models import *
from .forms import *
from .decorators import *

from datetime import *

def error404(request, *args, **kwargs):
    return render(request,'pages/404.html') 

def error500(request, *args, **kwargs):
    return render(request,'pages/500.html')
def export_movies_to_xlsx(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    student_queryset = Student.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-students.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Students'

    # Define the titles for columns
    columns = [
        'PK',
        'Name',
        'Phone number',
        'Email',
        'Active',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for student in student_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            student.pk,
            student.fullname,
            student.phonenumber,
            student.email,
            student.active,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
# Create your views here.
@unauthenticated_login
def loginPage(request):
        errorlogin = False
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, username=username, password=password)
            context={'user':user}
            if user is not None:
                login(request, user )
                return redirect('home')
            else:
                messages.info(request, 'Tên đăng nhập hoặc mật khẩu không đúng')
                errorlogin = True
                context={'user':user,'errorlogin':errorlogin}
                return render(request, 'pages/login.html', context)
        context={'errorlogin':errorlogin}
        return render(request, 'pages/login.html', context)

@unauthenticated_register
def registerPage(request):
        form = CreateUserForm()
        if request.method == 'POST':
            form = CreateUserForm(request.POST)
            if form.is_valid():
                user = form.save()  
                username = form.cleaned_data.get('username')

                user.is_superuser = 1
                user.is_staff = 1
                user.is_active = 1
                user.save()
                # group = Group.objects.get(name='manager')
                # user.groups.add(group)
                # group = Group.objects.get(name='admin')
                # user = form.save()
                messages.success(request, 'Tài khoản ' + username + ' tạo thành công')
                return redirect('login')
            else:
                messages.error(request, "Lỗi tạo tài khoản")
                return render(request, 'pages/register.html')
        context = {'form':form}
        return render(request, 'pages/register.html',context)
        # context = {'form':form}
        # return render(request, 'pages/register.html', context)

def logoutUser(request):
    logout(request)
    return redirect('login')
    
@login_required(login_url='login')
def home(request):
    notify = Notify.objects.all()
    listclass = Classname.objects.all()
    countstudent = Student.objects.filter(active=1).count()
    countteacher = Teacher.objects.filter(active=1).count()
    countclass = Classname.objects.filter(active=1).count()
    for i1 in listclass:
        startdate = datetime.strptime(str(i1.startdate), '%Y-%m-%d')
        today = datetime.now()
        student = Student.objects.all()
        if startdate <= today:
            if i1.active==0:
                # print(i1.pk) 
                i1.active=1
                i1.save()
        
        for a in student:
            if i1.pk==a.classname_id:
                if startdate <= today:
                    # print(a.classname_id) 
                    a.active=1
                    a.save()
                # if a.classname_id==i1.pk:
                #     a.active = 1
                #     a.save()
    schedule = Schedule.objects.all()
    for i2 in schedule:
        daylearn = datetime.strptime(str(i2.daylearn), '%Y-%m-%d')
        today = datetime.now()
        if daylearn < today:
            if i2.active==0:
                i2.active=1
                i2.save()
    for i3 in listclass:
        detailschedule = Schedule.objects.filter(id_classname = i3.pk).order_by('-daylearn')[0]
        print(detailschedule.active)
        daylearn = datetime.strptime(str(detailschedule.daylearn), '%Y-%m-%d')
        print(daylearn)
        studentinclass = Student.objects.filter(classname_id=i3.pk)
        today = datetime.now()
        if daylearn < today:
            i3.active=0
            i3.save()
            for i4 in studentinclass:
                print(i4.fullname)
                i4.active=0
            
        # for i4 in detailschedule:

        #     print(i4.daylearn)
    context = {'schedule':schedule,'notify':notify,'countstudent':countstudent,'countteacher':countteacher,'countclass':countclass}
    return render(request, 'pages/home.html',context)

# Class 
# Dat
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def listClass(request):
    # auto run 
    schedule = Schedule.objects.all()
    listclass = Classname.objects.all() 
    arraya = []
    for itema in listclass:
        SLa = {}
        SLa['idclass'] = itema.pk
        SLa['number'] = Student.objects.filter(classname_id=itema).count()
        arraya.append(SLa)
    arrayb = []
    for itemb in listclass:
        SLb = {}
        SLb['id'] = itemb.pk
        SLb['number'] = Schedule.objects.filter(id_classname=itemb.pk,active=1).count()
        arrayb.append(SLb)
    context = {'listclass': listclass,'schedule':schedule,'SLa':arraya,'SLb':arrayb}
    return render(request, 'class/listClass.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager'])
def createClass(request):
    form = CreateClassnameForm()
    unit = Unit.objects.all()
    area = Area.objects.all()
    room = Room.objects.all()
    timeshift = TimeShift.objects.all()
    timeweek = TimeWeek. objects.all()
    teacher = Teacher.objects.all()
    context = {'form':form, 'unit':unit, 'area':area, 'room':room, 'timeshift':timeshift, 'timeweek':timeweek, 'teacher':teacher}
    if request.method == 'POST':
        form = CreateClassnameForm(request.POST)
        if form.is_valid():
            idtimeshift = request.POST.get('timeshift')
            idtimeweek = request.POST.get('timeweek')
            idroom = request.POST.get('room')
            nameclass = request.POST.get('fullname')
            print(idtimeshift)
            print(idtimeweek)
            print(idroom)
            # if()
            form.save()
            count = 0
            start_date = datetime.strptime(str(request.POST.get('startdate')), '%Y-%m-%d')
            step = timedelta(days=1)
            time = TimeShift.objects.get(pk = request.POST.get('timeshift'))
            if int(request.POST.get('timeweek')) == 2:
                day = [1,3,5]
            elif int(request.POST.get('timeweek')) == 1:
                day = [0,2,4]
            while count < int(request.POST.get('datecount')):
                for i in day:
                    if i==start_date.date().weekday():
                        # kiem tra 
                        idclass = Classname.objects.get(timeshift = int(idtimeshift), timeweek = int(idtimeweek), room = int(idroom), fullname = str(nameclass))
                        print(start_date.date())
                        if idclass:
                            Schedule.objects.create(id_classname=int(idclass.pk), classname=str(request.POST.get('fullname')), daylearn = str(start_date.date()), timelearnstart = time.timestart, timelearnend = time.timeend, dayname="Buổi " + str(count+1),active=0)
                        count = count+1
                start_date += step
            # return redirect('createclass')
            return redirect('listclass')
            # check = Classname.objects.filter(timeweek = int(idtimeweek), room = int(idroom))
            # time = Timeshift.objects.filter()
            # check1 = Schedule.objects.filter()
            # print(check)
            # if check:
            #     print("ERROR")
            # else:
            #     form.save()
            #     count = 0
            #     start_date = datetime.strptime(str(request.POST.get('startdate')), '%Y-%m-%d')
            #     step = timedelta(days=1)
            #     time = TimeShift.objects.get(pk = request.POST.get('timeshift'))
            #     if int(request.POST.get('timeweek')) == 2:
            #         day = [1,3,5]
            #     else:
            #         day = [0,2,4]
            #     while count < int(request.POST.get('datecount')):
            #         for i in day:
            #             if i==start_date.date().weekday():
            #                 # kiem tra 
            #                 idclass = Classname.objects.get(timeshift = int(idtimeshift), timeweek = int(idtimeweek), room = int(idroom), fullname = str(nameclass))
            #                 print(start_date.date())
            #                 if idclass:
            #                     Schedule.objects.create(id_classname=int(idclass.pk), classname=str(request.POST.get('fullname')), daylearn = str(start_date.date()), timelearnstart = time.timestart, timelearnend = time.timeend, dayname="Buổi " + str(count+1),active=0)
            #                 count = count+1
            #         start_date += step
            #     # return redirect('createclass')
            #     return redirect('listclass') 
    return render(request, 'class/createClass.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def detailClass(request,pk):
    classname = Classname.objects.get(pk=pk)
    unit = Unit.objects.all()
    area = Area.objects.all()
    room = Room.objects.all()
    timeshift = TimeShift.objects.all()
    timeweek = TimeWeek. objects.all()
    teacher = Teacher.objects.all()
    form = UpdateClassForm(instance=classname)
    if request.method == 'POST':
        form = UpdateClassForm(request.POST, instance=classname)
        if form.is_valid():
            form.save()
            return redirect('listclass')
    context={'form':form, 'classname':classname, 'unit':unit, 'area':area, 'room':room, 'timeshift':timeshift, 'timeweek':timeweek, 'teacher':teacher}
    return render(request,'class/detailClass.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def detailClassStudent(request,pk):
    pk = pk
    classname = Classname.objects.get(pk=pk)
    liststudentinclass = Student.objects.all()
    timeshift = TimeShift.objects.get(pk=classname.timeshift_id)
    print(timeshift.timestart)
    schedule = Schedule.objects.all()
    context={'timeshift':timeshift,'liststudentinclass':liststudentinclass,'pk':pk, 'classname':classname, 'schedule':schedule}
    return render(request,'class/detailClassStudent.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def checkinClassStudent(request,pk):
    pk = pk
    classname = Classname.objects.get(pk=pk)
    liststudentinclass = Student.objects.all()
    timeshift = TimeShift.objects.get(pk=classname.timeshift_id)
    print(timeshift.timestart)
    checkinclass = CheckInClass.objects.all().order_by("daylearn")
    if request.method=='POST':
        
        checkin = request.POST.getlist('check')
        for id in checkin:
            print(int(id))
            check = CheckInClass.objects.get(pk=int(id))
            if check.active == 0:
                check.active = 1
            check.save() 

        return redirect('listclass')
    schedule = Schedule.objects.all().order_by("daylearn")
    today = datetime.now().date()
    print('hom nay ',today)
    # schedule1 = Schedule.objects.fillter(id_classname=pk).count()
    # print(schedule1)
    context={'checkinclass':checkinclass,'today':today,'timeshift':timeshift,'liststudentinclass':liststudentinclass,'pk':pk, 'classname':classname, 'schedule':schedule}
    return render(request,'class/checkinClassStudent.html',context)

# Student 
# Dat
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','student','teacher'])
def listStudent(request):
    form = CreateStudentForm()
    liststudent = Student.objects.all()
    listclass = Classname.objects.all()
    for i1 in listclass:
        startdate = datetime.strptime(str(i1.startdate), '%Y-%m-%d')
        today = datetime.now()
        # student = Student.objects.all()
        if startdate <= today:
            # print(i1.pk) 
            i1.active=1
            i1.save()
        for a in liststudent:
            if i1.pk==a.classname_id:
                if startdate <= today:
                    # print(a.classname_id) 
                    a.active=1
                    a.save()
    context = {'listclass': listclass,'liststudent': liststudent}
    return render(request, 'student/student.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def checkinStudent(request,pk):
    student = Student.objects.get(pk=pk) 
    classrequest = Classname.objects.get(pk=student.classname_id)
    time = TimeShift.objects.get(pk = classrequest.timeshift_id)
    checkinclass = CheckInClass.objects.filter(id_student=student.pk)
    context = {'checkinclass':checkinclass, 'classrequest':classrequest,'time':time,'student':student}
    if request.method == 'POST':
        if student.havedetailschedule == 0:
            student.havedetailschedule = 1
            student.save()
            # checkinclass = CheckInClass.objects.filter(id_student=student.pk) 
            count = 0
            start_date = datetime.strptime(str(classrequest.startdate), '%Y-%m-%d')
            step = timedelta(days=1)
            # print(classrequest.timeweek_id)
            if int(classrequest.timeweek_id) == 1:
                day = [0,2,4]
            elif int(classrequest.timeweek_id) == 2:
                day = [1,3,5]
            while count < int(classrequest.datecount):
                for i in day:
                    if i==start_date.date().weekday():
                        # print(start_date.date())
                        CheckInClass.objects.create(id_student=pk,student=str(student.fullname),classname=str(classrequest.fullname), daylearn = str(start_date.date()), timelearnstart = time.timestart, timelearnend = time.timeend, dayname="Buổi " + str(count+1),active=0)
                        count = count+1
                start_date += step
            # context = {'checkinclass':checkinclass, 'classrequest':classrequest,'time':time,'student':student}
            # return render(request, 'student/checkinStudent.html', context)
            return redirect('liststudent')
        else:
            student.havedetailschedule = 0
            student.active = 0
            student.save()
            CheckInClass.objects.filter(id_student=student.pk).delete()
            return redirect('liststudent')
    # context = {'checkinclass':checkinclass, 'classrequest':classrequest,'time':time,'student':student}
    return render(request, 'student/checkinStudent.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','student'])
def createStudent(request):
    form = CreateStudentForm()
    gender = Gender.objects.all()
    unit = Unit.objects.all()
    today = date.today()
    error=0
    classname = Classname.objects.all()
    if request.method == 'POST':
        form = CreateStudentForm(request.POST)
        # print('Print Form: ', form.is_valid())
        if form.is_valid(): 
            form.save()
            idstudent = Student.objects.filter().order_by("-pk")[0].pk
            # print(idstudent)
            refee = Student.objects.get(pk=idstudent)
            fee = int(request.POST.get('fee'))
            idunit = int(request.POST.get('unit'))
            feeunit = int(Unit.objects.get(pk=idunit).fee)
            if feeunit>=fee:
                refee.fee_remain = feeunit-fee
            else:
                refee.fee_remain = 0
            refee.save()
            send_mail(
                subject = 'XÁC NHẬN ĐĂNG KÝ HỌC VIÊN', # title mail
                message = 'Chào '+str(request.POST.get('fullname')) + ',\nBạn vừa hoàn thành đăng kí học viên tại HITECH.\nLớp '+str(refee.classname)+' hiện tại chưa bắt đầu học.\nTrung tâm sẽ gửi thời khóa biểu cho bạn sớm nhất.', # nội dung mail
                from_email= None, # tài khoản
                auth_password= None, # mk
                recipient_list = [form.cleaned_data.get('email')],# mail người nhận
                fail_silently = False,
            )
            # return redirect('createstudent')
            return redirect('liststudent')
        else:
            error=1
    context = {'error':error, 'gender':gender, 'today':today, 'unit':unit, 'classname':classname,'form':form}
    return render(request, 'student/createStudent.html', context)

# cái này chưa xong
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def detailStudent(request,pk):
    student = Student.objects.get(pk=pk)
    classname = Classname.objects.all()
    unit = Unit.objects.all()
    schedule = Schedule.objects.all()
    feeremain = student.fee_remain
    error=0
    if request.method == 'POST':
        # print(request.POST.get('classname'))
        if request.POST.get('classname') is not None:
            student.classname_id = request.POST.get('classname')
            print(request.POST.get('classname'))
            student.email = request.POST.get('email')
            student.phonenumber = request.POST.get('phonenumber')
            student.adress = request.POST.get('adress')
            student.phonenumber_family = request.POST.get('phonenumber_family')
            # chua xong
            # tong hoc phi nhap vao va hoc phi da dong
            if request.POST.get('fee') == '':
                feein = 0
            else:
                feein = int(request.POST.get('fee'))
            sumfee = feein + student.fee
            # -------------------------------------
            # id va hoc phi khoa hoc
            idunit = int(request.POST.get('unit'))
            # print('id unit: ',idunit)
            feeunit = int(Unit.objects.get(pk=idunit).fee)
            # print('fee unit: ',feeunit)
            # -------------------------------------
            if sumfee > feeunit:
                student.fee = feeunit
                student.fee_remain = 0
            else:
                student.fee = sumfee
                student.fee_remain = feeunit - sumfee
            # print('tien con thieu: ',student.fee_remain)
            # print('tien da dong: ',student.fee)
            # -------------------------------------
            student.note = request.POST.get('note')
            if request.POST.get('active') is not None:
                student.active = 1
            else:
                student.active = 0
            student.save()
            send_mail(
                subject = 'XÁC NHẬN ĐÃ THAY ĐỔI THÔNG TIN HỌC VIÊN', # title mail
                message = 'Chào '+ str(request.POST.get('fullname')) + 
                            ',\nThông tin học viên HITECH của bạn vừa được cập nhật. \nVui lòng trả lời tin mail này để xác nhận.', # nội dung mail
                from_email= None, # tài khoản
                auth_password= None, # mk
                recipient_list = [request.POST.get('email')],# mail người nhận
                fail_silently = False,
            )
            return redirect('liststudent')
        else:
            error=1
    context = {'error':error, 'feeremain':feeremain, 'schedule':schedule, 'student':student, 'classname': classname, 'unit': unit}
    return render(request,'student/detailStudent.html',context)

# Teacher 
# Dat
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def listTeacher(request):
    listteacher = Teacher.objects.all()
    classname = Classname.objects.all()
    array = []
    for item in listteacher:
        SL = {}
        SL['teachername'] = item.pk
        SL['number'] = Classname.objects.filter(teacher_id=item).count()
        array.append(SL)
        # print(array)
    context = {'classname': classname,'listteacher': listteacher,'SL':array}
    return render(request, 'teacher/teacher.html', context)
    
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def createTeacher(request):
    gender = Gender.objects.all()
    listteacher = Teacher.objects.all()
    form = CreateTeacherForm()
    if request.method == 'POST':
        form = CreateTeacherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listteacher')
    context = {'form':form, 'listteacher': listteacher, 'gender':gender}
    return render(request, 'teacher/createTeacher.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def detailTeacher(request,pk):
    teacher = Teacher.objects.get(pk=pk)
    gender = Gender.objects.all()
    form = UpdateTeacherForm(instance=teacher)
    if request.method == 'POST':
        form = UpdateTeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            return redirect('listteacher')
    context = {'form':form, 'teacher':teacher, 'gender': gender}
    return render(request,'teacher/detailTeacher.html',context)

# Manager 
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff','teacher'])
def editProfile(request,pk):
    detailuser = User.objects.get(pk=pk)
    if request.method == 'POST':
        detailuser.last_name = request.POST.get('last_name')
        detailuser.first_name = request.POST.get('first_name')
        detailuser.save()
        return redirect('listuser')
    context={'detailuser':detailuser} 
    return render(request,'manager/editProfile.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def newRoom(request):
    form = CreateRoomForm()
    area = Area.objects.all()
    context={'form':form,'area':area} 
    if request.method == 'POST':
        form = CreateRoomForm(request.POST)
        if form.is_valid(): 
            form.save()
            return render(request,'manager/newRoom.html',context)
    return render(request,'manager/newRoom.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editRoom(request,pk):
    room = Room.objects.get(pk=pk)
    area = Area.objects.all()
    context={'area':area,'room':room} 
    if request.method == 'POST':
        room.fullname = request.POST.get('fullname')
        room.area_id = request.POST.get('area')
        if request.POST.get('active') == 'on':
            room.active = 1
        else:
            room.active = 0
        room.save()
        return redirect('/listarea/')
    return render(request,'manager/editRoom.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def newArea(request):
    form = CreateAreaForm()
    context={'form':form} 
    if request.method == 'POST':
        form = CreateAreaForm(request.POST)
        if form.is_valid(): 
            form.save()
            return redirect('newroom')
    return render(request,'manager/newArea.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editArea(request,pk):
    area = Area.objects.get(pk=pk)
    room = Room.objects.all()
    if request.method == 'POST':
        area.fullname = str(request.POST.get('fullname'))
        area.phonenumber = str(request.POST.get('phonenumber'))
        area.email = str(request.POST.get('email'))
        area.adress = str(request.POST.get('adress'))
        area.note = str(request.POST.get('note'))
        if request.POST.get('active') == 'on':
            area.active = 1
        else:
            area.active = 0
        area.save()
        return redirect('listarea')
    context={'area':area,'room':room}         
    return render(request,'manager/editArea.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def listArea(request):
    form = CreateAreaForm()
    listarea = Area.objects.all()
    context={'form':form,'listarea':listarea} 
    return render(request,'manager/listArea.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def notify(request):
    error = 0
    notify = Notify.objects.all()
    if request.method == 'POST':
        if request.POST.get('title') != '':
            title = request.POST.get('title')
            content = request.POST.get('content')
            if request.POST.get('active') is None:
                active = 1
            else:
                active = 0
            Notify.objects.create(title=title,content=content,active=active)
            error = 0
        else:
            error = 1
    context={'notify':notify,'error':error} 
    return render(request,'manager/notify.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editNotify(request,pk):
    noti = Notify.objects.get(pk=pk)
    if request.method == 'POST':
        if request.POST.get('active') == 'on':
            noti.active = 0
        else:
            noti.active = 1
        noti.content = request.POST.get('content')
        noti.save()
        print(noti.content)
        return redirect('notify')
    context={'noti':noti} 
    return render(request,'manager/editNotify.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def createTimeShift(request):
    form = CreateTimeShiftForm()
    timeshift = TimeShift.objects.all()
    if request.method == 'POST':
        form = CreateTimeShiftForm(request.POST)
        if form.is_valid(): 
            form.save()
    context={'form':form,'timeshift':timeshift} 
    return render(request,'manager/createTimeShift.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editTimeShift(request,pk):
    timeshift = TimeShift.objects.get(pk=pk)
    if request.method == 'POST':
        timeshift.fullname = request.POST.get('fullname')
        timeshift.infomation = request.POST.get('infomation')
        timeshift.timestart = request.POST.get('timestart')
        timeshift.timeend = request.POST.get('timeend')
        if request.POST.get('active') == 'on':
            timeshift.active = 1
        else:
            timeshift.active = 0
        timeshift.save()
        return redirect('createtimeshift')
    context={'timeshift':timeshift} 
    return render(request,'manager/editTimeShift.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def createUnit(request):
    form = CreateUnitForm()
    unit = Unit.objects.all()
    if request.method == 'POST':
        form = CreateUnitForm(request.POST)
        if form.is_valid(): 
            form.save()
    context={'form':form,'unit':unit} 
    return render(request,'manager/createUnit.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editInformationUnit(request,pk):
    unit = Unit.objects.get(pk=pk)
    if request.method == 'POST':
        unit.infomation = request.POST.get('infomation')
        unit.fee = request.POST.get('fee')
        if request.POST.get('active') == 'on':
            unit.active = 1
        else:
            unit.active = 0
        unit.save()
        return redirect('createunit')
    context={'unit':unit} 
    return render(request,'manager/editInformationUnit.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager'])
def listUser(request):
    listuser = User.objects.all()
    array = []
    for item in listuser:
        USERGROUP = {}
        USERGROUP['iduser'] = item.pk
        USERGROUP['groupuser'] = str(item.groups.all()[0].name)
        array.append(USERGROUP)
    context={'listuser':listuser,'USERGROUP':array}
    return render(request,'manager/listUser.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager'])
def createStafAccount(request):
    form = CreateUserForm()
    context = {'form':form}
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            # user.is_superuser = 1
            # user.is_staff = 1
            user = form.save()  
            # print(request.POST.get('role'))
            username = form.cleaned_data.get('username')
            if int(request.POST.get('role')) == 1:
                user.is_superuser = 1
                user.is_staff = 1
                user.is_active = 1
                group = Group.objects.get(name='admin')
                user.groups.add(group)
            elif int(request.POST.get('role')) == 2:
                user.is_staff = 1
                user.is_active = 1
                group = Group.objects.get(name='manager')
                user.groups.add(group)
            elif int(request.POST.get('role')) == 3:
                user.is_active = 1
                group = Group.objects.get(name='staff')
                user.groups.add(group)
            elif int(request.POST.get('role')) == 4:
                user.is_active = 1
                group = Group.objects.get(name='teacher')
                user.groups.add(group)
            elif int(request.POST.get('role')) == 5:
                user.is_active = 1
                group = Group.objects.get(name='student')
                user.groups.add(group)
            user.save()
            messages.success(request, 'Tài khoản ' + username + ' tạo thành công')
            return redirect('listuser') 
        else:
            messages.error(request, "Lỗi tạo tài khoản")
            return render(request,'manager/createStafAccount.html',context)
    return render(request,'manager/createStafAccount.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager'])
def updateUser(request,pk):
    detailuser = User.objects.get(pk=pk)
    group = detailuser.groups.all()[0].name 
    # for i in group: 
    print(group) 
    rolename = str(group)
    if request.method == 'POST':
        detailuser.groups.clear()
        if str(request.POST.get('role')) == 'admin':
            is_superuser = 1
            is_staff = 1
            is_active = 1
            group = Group.objects.get(name='admin')
            detailuser.groups.add(group)
        elif str(request.POST.get('role')) == 'manager':
            is_superuser = 0
            is_staff = 1
            is_active = 1
            group = Group.objects.get(name='manager')
            detailuser.groups.add(group)
        elif str(request.POST.get('role')) == 'staff':
            is_superuser = 0
            is_staff = 0
            is_active = 1
            group = Group.objects.get(name='staff')
            detailuser.groups.add(group)
        elif str(request.POST.get('role')) == 'teacher':
            is_superuser = 0
            is_staff = 0
            is_active = 1
            group = Group.objects.get(name='teacher')
            detailuser.groups.add(group)
        elif str(request.POST.get('role')) == 'student':
            is_superuser = 0
            is_staff = 0
            is_active = 1
            group = Group.objects.get(name='student')
            detailuser.groups.add(group)
        # print(is_active)
        # print(is_staff)
        # print(is_superuser)
        detailuser.is_active = is_active
        detailuser.is_staff = is_staff
        detailuser.is_superuser = is_superuser
        detailuser.save()
        # if request.POST.get('is_superuser'):
        #     group = Group.objects.get(name='admin')
        #     detailuser.groups.add(group)
        # elif request.POST.get('is_staff'):
        #     group = Group.objects.get(name='manager')
        #     detailuser.groups.add(group)
        # else:
        #     group = Group.objects.get(name='staff')
        #     detailuser.groups.add(group)
        return redirect('listuser')
    context={'detailuser':detailuser,'rolename':rolename}
    return render(request,'manager/updateUser.html',context)

# Contact 
@login_required(login_url='login')
def contact(request):
    area = Area.objects.first()
    context = {'area':area}
    return render(request,'contact/contact.html',context)
 
# RoleError 
@login_required(login_url='login')
def errorRole(request):
    return render(request,'pages/errorRole.html')
  